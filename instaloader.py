import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import os

# تنظیمات مرورگر
chrome_options = Options()
chrome_options.add_argument("--start-maximized")

# ⚙️ تنظیم مسیر دانلود (اختیاری - اگر بخوای دانلودها به جای دسکتاپ، جای خاصی بری)
download_dir = r"F:\Downloaded_Videos"  # مسیر دلخواهت رو اینجا وارد کن
os.makedirs(download_dir, exist_ok=True)

# تنظیمات دانلود برای Chrome
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
chrome_options.add_experimental_option("prefs", prefs)

# راه‌اندازی مرورگر
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# خواندن فایل اکسل
excel_path = r"F:\sourcecode\Instaloader\files\insta-downloader.xlsx"
try:
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
except Exception as e:
    print(f"خطا در باز کردن فایل اکسل: {e}")
    exit()

# پیدا کردن ستون "URL" در سطر اول
url_column = None
for cell in sheet[1]:  # سطر اول (هدر)
    if cell.value == "URL":
        url_column = cell.column
        break

if not url_column:
    print('❌ ستونی با عنوان "URL" در سطر اول پیدا نشد.')
    exit()

# لیست تمام لینک‌ها از سطر 2 به بعد
url_rows = []
for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=url_column).value
    if cell_value and str(cell_value).strip().startswith("http"):
        url_rows.append((row, cell_value))

if not url_rows:
    print("❌ هیچ لینک معتبری در ستون URL پیدا نشد.")
    exit()

print(f"✅ {len(url_rows)} لینک یافت شد. شروع فرآیند دانلود...")

# باز کردن صفحه اصلی
driver.get("https://en1.savefrom.net/14KS/")

# منتظر بمان تا صفحه آماده باشد
wait = WebDriverWait(driver, 20)

for row_idx, video_url in url_rows:
    try:
        print(f"\n📥 پردازش لینک سطر {row_idx}: {video_url}")

        # پاک کردن و وارد کردن لینک جدید
        input_box = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='sf_url']")))
        input_box.clear()
        input_box.send_keys(video_url)

        # کلیک روی دکمه دانلود
        submit_btn = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
        submit_btn.click()
        print("لینک ارسال شد. صبر برای ظاهر شدن گزینه‌های دانلود...")

        # صبر 10 ثانیه (یا تا لینک دانلود ظاهر بشه)
        download_link = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'link-download') and contains(@href, 'get?')]"))
        )

        # گرفتن نام فایل از ویژگی `download` (اختیاری - فقط برای نمایش)
        filename = download_link.get_attribute("download")
        print(f"✅ لینک دانلود پیدا شد: {filename}")

        # کلیک روی لینک دانلود
        download_link.click()
        print("📥 دانلود شروع شد...")

        # صبر برای شروع دانلود (تا فایل دانلود شروع بشه قبل از رفتن به لینک بعدی)
        time.sleep(8)  # اگر ویدیو بزرگه، این زمان رو افزایش بده

        # بازگشت به صفحه اصلی برای لینک بعدی
        driver.get("https://en1.savefrom.net/14KS/")
        time.sleep(3)  # کمی صبر برای بارگذاری مجدد

    except TimeoutException:
        print(f"❌ زمان‌سررسید! لینک دانلود برای {video_url} پیدا نشد (ممکن است مشکل اینترنت یا ضدربات باشد).")
        driver.get("https://en1.savefrom.net/14KS/")  # ریست صفحه
        time.sleep(5)

    except Exception as e:
        print(f"❌ خطای ناشناخته برای لینک سطر {row_idx}: {str(e)}")
        driver.get("https://en1.savefrom.net/14KS/")
        time.sleep(5)

# پایان
print("\n🎉 تمام لینک‌ها پردازش شدند.")
input("برای بستن مرورگر، Enter بزنید...")
driver.quit()